"""The plain-English description of each parameter, from item_desc.xlsx.

The reply side cannot match on headings - taxpayers write "Query No:2" or
nothing at all - and a bare parameter title is thin evidence for the model to
decide whether a paragraph answers it. The descriptions say what the defect
actually alleges ("ITC claimed in GSTR-3B/GSTR-9 Table 6A is higher than the
category-wise ITC in Table 6I"), which is what a reply argues against.

The spreadsheet numbers its rows 1-22 in the department's own order, which is
not the A/B/C order of Parameters_TN.pdf, and row 16 (interest under s.50(1))
is not one of the 21 parameters at all. The mapping below is therefore explicit
rather than positional, and is asserted against params.py on import.

The descriptions are given to the model as *context for judging*, never as text
to reproduce. reply_llm.py checks every answer against the reply document
itself, so a sentence lifted from a description cannot survive.
"""

from pathlib import Path

from openpyxl import load_workbook

from paths import ROOT
from params import REPORTED, TITLE

SOURCE = ROOT / "item_desc.xlsx"

# item_desc.xlsx S.no -> parameter id. Row 16 ("Interest under Section 50(1)")
# describes interest generally and has no sheet of its own.
ROW_TO_PID = {
    1: "A1", 2: "A2", 3: "A3", 4: "A4", 5: "A5", 6: "A6",
    7: "B1", 8: "B2", 9: "B3", 10: "B4", 11: "B5", 12: "B6",
    13: "B7", 14: "B8", 15: "B9",
    16: None,
    17: "B10",
    18: "C3", 19: "C4", 20: "C5", 21: "C1", 22: "C2",
}


def _load():
    if not SOURCE.exists():
        return {}
    ws = load_workbook(SOURCE, read_only=True)["Sheet1"]
    out = {}
    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        sno, item, desc = row
        try:
            n = int(str(sno).strip())
        except (TypeError, ValueError):
            continue                      # the title row and the blank one
        pid = ROW_TO_PID.get(n)
        if pid and desc:
            out[pid] = " ".join(str(desc).split())
    return out


DESC = _load()

# A parameter with no description still works - it falls back to its title -
# but a *wrong* id in the mapping would silently mis-describe a whole sheet.
assert set(DESC) <= set(REPORTED), f"unknown ids in mapping: {set(DESC) - set(REPORTED)}"


def brief(pid):
    """What to tell the model this parameter alleges."""
    d = DESC.get(pid)
    return f"{TITLE[pid]}\n     {d}" if d else TITLE[pid]


if __name__ == "__main__":
    missing = [p for p in REPORTED if p not in DESC]
    print(f"{len(DESC)} of {len(REPORTED)} parameters described"
          + (f"; no description for {missing}" if missing else ""))
    for pid in REPORTED:
        print(f"\n{pid}  {brief(pid)}")
