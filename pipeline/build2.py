"""Stage 5 - the workbook. One sheet per parameter, one row per company.

    GSTIN | Trade name | Notice (SCN) defect | Taxpayer reply | Officer finding

The notice is the spine: a company appears on a sheet only if its own DRC-01
raised that parameter, so there is no row without a defect. The reply column is
what the model found, or "No reply". The finding column is deliberately empty -
the orders are not being read yet.

Monospace on the three text columns, because the notice tables only hold their
shape while the column padding pdftotext produced is preserved. Tables the model
returned as rows are re-padded the same way.

    python3 build2.py            -> register_21-22.xlsx
    python3 build2.py <name>     -> <name>.xlsx
"""

import json
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from paths import WORK, OUT_NAME, OUT_DIR
from params import PARAMS, SHEET_NAME, TITLE

NO_REPLY = "No reply"
NO_FINDING = "No finding"

MONO = Font(name="Menlo", size=9)
HEAD_FONT = Font(name="Helvetica Neue", size=10, bold=True, color="FFFFFF")
META_FONT = Font(name="Helvetica Neue", size=10)
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")
THIN = Side(style="thin", color="BFC7D5")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)

COLS = [("GSTIN", 20), ("Trade name", 30),
        ("Notice (SCN) defect", 95), ("Taxpayer reply", 85),
        ("Verdict", 16), ("Officer finding", 80)]

MAX_CELL = 30000        # Excel's own limit is 32767
MAX_ROW_HEIGHT = 409    # Excel's own limit

# A scanned PDF with no text layer sometimes yields control bytes rather than
# nothing, and Excel refuses to store them at all - one such cell aborts the
# whole save.
ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037\177]")


def safe(text):
    return ILLEGAL.sub("", str(text or ""))


def render_table(t):
    """A model table back into padded columns, so it lines up under Menlo."""
    rows = [t["headers"]] + t["rows"]
    width = [max(len(str(r[i])) for r in rows) for i in range(len(t["headers"]))]
    out = []
    if t.get("title"):
        out.append(t["title"])
    for n, r in enumerate(rows):
        out.append("  ".join(str(c).ljust(width[i]) for i, c in enumerate(r)).rstrip())
        if n == 0:
            out.append("  ".join("-" * w for w in width))
    return "\n".join(out)


def notice_cell(sec, fixed):
    """The notice section: repaired tables when the repair was proved, else raw.

    pdftotext wraps the department's wide tables, so a figure can arrive split
    by a space or continued on the line below. notice_tables.py rebuilds those
    and checks every figure against the section's own digits; a section whose
    repair did not pass keeps the literal text, which is what the register has
    always shown.
    """
    if fixed and fixed.get("verified"):
        parts = [fixed["prose"]] if fixed.get("prose") else []
        parts += [render_table(t) for t in fixed.get("tables") or []]
        joined = "\n\n".join(p for p in parts if p).strip()
        if joined:
            return joined[:MAX_CELL]
    return sec["text"][:MAX_CELL]


def reply_cell(rec):
    if not rec or rec.get("no_reply"):
        return NO_REPLY
    text = rec["text"]
    if not rec.get("verified", True) and rec.get("fallback_text"):
        # The figures the model wrote were not all in the document, so the cell
        # becomes the literal lines it pointed at instead.
        return rec["fallback_text"][:MAX_CELL]
    parts = [text] if text else []
    parts += [render_table(t) for t in rec.get("tables") or []]
    return ("\n\n".join(p for p in parts if p).strip() or NO_REPLY)[:MAX_CELL]


def order_cell(rec):
    """The officer's finding, and separately the verdict it supports."""
    if not rec or rec.get("no_finding"):
        return "", NO_FINDING
    verdict = rec.get("verdict") or "unclear"
    if not rec.get("verified", True) and rec.get("fallback_text"):
        return verdict, rec["fallback_text"][:MAX_CELL]
    parts = [rec["text"]] if rec.get("text") else []
    parts += [render_table(t) for t in rec.get("tables") or []]
    body = "\n\n".join(p for p in parts if p).strip()
    return (verdict, body[:MAX_CELL]) if body else ("", NO_FINDING)


def est_height(*cells):
    """Roughly how tall the row needs to be, capped."""
    lines = 0
    for c, width in zip(cells, (95, 85, 80)):
        for line in str(c).splitlines() or [""]:
            lines += max(1, -(-len(line) // width))
    return min(MAX_ROW_HEIGHT, max(30, lines * 11.5))


def add_sheet(wb, pid, rows):
    ws = wb.create_sheet(SHEET_NAME[pid][:31])
    ws.append([f"{pid}  -  {TITLE[pid]}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(row=1, column=1)
    c.font = Font(name="Helvetica Neue", size=12, bold=True, color="1F3864")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append([h for h, _ in COLS])
    for i in range(1, len(COLS) + 1):
        h = ws.cell(row=2, column=i)
        h.font, h.fill, h.border = HEAD_FONT, HEAD_FILL, BOX
        h.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    for n, r in enumerate(rows):
        ws.append([r["gstin"], safe(r["trade_name"]), safe(r["scn"]),
                   safe(r["reply"]), r["verdict"], safe(r["order"])])
        row = ws.max_row
        for i in range(1, len(COLS) + 1):
            cell = ws.cell(row=row, column=i)
            cell.alignment = TOP_WRAP
            cell.border = BOX
            cell.font = MONO if i >= 3 else META_FONT
            if n % 2:
                cell.fill = ALT_FILL
        ws.row_dimensions[row].height = est_height(r["scn"], r["reply"],
                                                   r["order"])

    for i, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    return ws


def contents(wb, tally, total_rows):
    ws = wb.create_sheet("Contents", 0)
    ws.append(["GSTR-9 scrutiny register - FY 2021-22"])
    ws.cell(row=1, column=1).font = Font(name="Helvetica Neue", size=14, bold=True,
                                         color="1F3864")
    ws.append([])
    ws.append(["Sheet", "Parameter", "Companies", "Replies", "No reply",
               "Findings", "No finding"])
    for i in range(1, 8):
        h = ws.cell(row=3, column=i)
        h.font, h.fill, h.border = HEAD_FONT, HEAD_FILL, BOX
        h.alignment = Alignment(vertical="center", horizontal="center")
    for pid, _, title, _ in PARAMS:
        n, answered, decided = tally.get(pid, (0, 0, 0))
        ws.append([SHEET_NAME[pid], title, n, answered, n - answered,
                   decided, n - decided])
        for i in range(1, 8):
            ws.cell(row=ws.max_row, column=i).font = META_FONT
    ws.append([])
    ws.append(["", "Total rows", total_rows])
    ws.cell(row=ws.max_row, column=2).font = Font(name="Helvetica Neue", size=10,
                                                  bold=True)
    for i, w in enumerate((26, 60, 12, 12, 12, 12, 12), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def main():
    name = sys.argv[1] if len(sys.argv) > 1 else OUT_NAME
    scn = json.loads((WORK / "scn.json").read_text())
    rpath = WORK / "reply.json"
    reply = json.loads(rpath.read_text()) if rpath.exists() else {}
    opath = WORK / "order.json"
    order = json.loads(opath.read_text()) if opath.exists() else {}
    fpath = WORK / "notice_fixed.json"
    fixed = json.loads(fpath.read_text()) if fpath.exists() else {}

    wb = Workbook()
    wb.remove(wb.active)

    tally, total = {}, 0
    for pid, _, _, _ in PARAMS:
        rows = []
        for gstin in sorted(scn):
            sec = scn[gstin]["params"].get(pid)
            if not sec:
                continue                      # no notice defect, no row
            rec = (reply.get(gstin, {}).get("params") or {}).get(pid)
            orec = (order.get(gstin, {}).get("params") or {}).get(pid)
            verdict, finding = order_cell(orec)
            rows.append({"gstin": gstin,
                         "trade_name": scn[gstin]["trade_name"],
                         "scn": notice_cell(sec, (fixed.get(gstin) or {}).get(pid)),
                         "reply": reply_cell(rec),
                         "verdict": verdict,
                         "order": finding})
        # A parameter no notice in this corpus raised still gets its sheet, so
        # the workbook always has all 21 - an empty sheet says "not raised",
        # a missing sheet looks like a pipeline that lost it.
        ws = add_sheet(wb, pid, rows)
        if not rows:
            ws.cell(row=3, column=1,
                    value="No notice in this dataset raised this parameter.")
            ws.cell(row=3, column=1).font = META_FONT
            tally[pid] = (0, 0, 0)
            continue
        answered = sum(1 for r in rows if r["reply"] != NO_REPLY)
        decided = sum(1 for r in rows if r["order"] != NO_FINDING)
        tally[pid] = (len(rows), answered, decided)
        total += len(rows)

    contents(wb, tally, total)
    out = OUT_DIR / f"{name}.xlsx"
    wb.save(out)

    ans = sum(a for _, a, _ in tally.values())
    dec = sum(d for _, _, d in tally.values())
    print(f"{out.name}: {len(tally)} sheets, {total} rows, "
          f"{ans} with a reply, {total - ans} '{NO_REPLY}', "
          f"{dec} with a finding")
    for pid, _, title, _ in PARAMS:
        if pid in tally:
            n, a = tally[pid]
            print(f"  {pid:<4} {n:>3} rows  {a:>3} replies   {title}")


if __name__ == "__main__":
    main()
